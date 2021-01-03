from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import load_workbook
import sqlalchemy as db


#selenium으로 접속하는 방법
def into_selenium(url):
    #브라우져 옵션 설정
    options = Options()
    #화면 안뜨게
    options.headless = True
    #유저 에이전트 설정
    #Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36
    options.add_argument("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36")
    broswer = webdriver.Chrome("./find_stock/chromedriver.exe",options = options)
    # broswer.maximize_window()
    #브라우저 연결
    broswer.get(url)
    return broswer
class stock_score:
    def __init__(self):
        

if __name__ == '__main__':
    # #엑셀파일에서 종목 코드 찾기
    # wb = load_workbook("./find_stock/stock_codenum.xlsm",data_only = True)
    # ws = wb.active
    # stock_code = ""
    # for stocknum in ws.iter_rows(min_row=2):
    #     if "ESR켄달스퀘어리츠" == stocknum[0].value:
    #         stock_code = str(stocknum[4].value)
    #         break
    #데이터 베이스 연결 설정
    sqlalchemy_database_url = 'mysql+pymysql://root:1234@localhost:3306/stockproject'
    engine = db.create_engine(sqlalchemy_database_url,echo=False)
    connection = engine.connect()

    # 데이터 베이스 테이블 결정
    metadata = db.MetaData()
    table = db.Table('stock_info',metadata,autoload=True,autoload_with=engine)

    query = db.select([table]).where(table.columns.stock_type == "KOSDAQ")

    result_proxy = connection.execute(query)
    result_set = result_proxy.fetchall()
    for stock_info in result_set:
        try:
            broswer = into_selenium(f"https://navercomp.wisereport.co.kr/v2/company/c1030001.aspx?cmp_cd={stock_info[1]}&cn=")
            #손익 계산서
            broswer.find_element_by_id("rpt_tab1").click()
            #매출액,매출 총이익, 영업이익, 당기순이익
            total_sales = []
            profit_sales = []
            operating_profit = []
            net_profit = []
            for i in range(2,7):
                total_sales.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[1]/td[{i}]').text).replace(',',''))
                profit_sales.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[26]/td[{i}]').text).replace(',',''))
                operating_profit.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[58]/td[{i}]').text).replace(',',''))
                net_profit.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[218]/td[{i}]').text).replace(',',''))

            #재무 상태표
            broswer.find_element_by_id("rpt_tab2").click()
            time.sleep(0.5)
            #자산 총계, 유동 자산, 현금 및 현금성 자산, 총 부채, 유동 부채,이익 잉여금
            total_assets = []
            liquid_assets = []
            cash_assets = []
            total_debt = []
            current_liabilities = []
            surplus_profit = []
            for i in range(2,7):
                total_assets.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[1]/td[{i}]').text).replace(',',''))
                liquid_assets.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[2]/td[{i}]').text).replace(',',''))
                cash_assets.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[47]/td[{i}]').text).replace(',',''))
                total_debt.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[127]/td[{i}]').text).replace(',',''))
                current_liabilities.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[127]/td[{i}]').text).replace(',',''))
                surplus_profit.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[243]/td[{i}]').text).replace(',',''))

            #현금 흐름표
            broswer.find_element_by_id("rpt_tab3").click()
            time.sleep(0.5)
            #영업활동으로 인한 현금흐름, 투자활동으로 인한 현금 흐름, 재무활동으로 인한 현금 흐름
            business_activities = []
            investment_activities = []
            financial_activities = []
            for i in range(2,7):
                business_activities.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[1]/td[{i}]').text).replace(',',''))
                investment_activities.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[144]/td[{i}]').text).replace(',',''))
                financial_activities.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[242]/td[{i}]').text).replace(',',''))

            broswer.get(f"https://navercomp.wisereport.co.kr/v2/company/c1040001.aspx?cmp_cd={stock_info[1]}&cn=")
            #EPS, PER, EV_EBITDA, PSR, PBR, INDUSTRIES_AVG_PER
            EPS = []
            PER = []
            EV_EBITDA = []
            PSR = []
            PBR = []
            INDUSTRIES_AVG_PER = (broswer.find_element_by_xpath('/html/body/div/form/div[1]/div/div[2]/div[1]/div/table/tbody/tr[3]/td/dl/dt[4]/b').text).replace(',','')
            for i in range(2,7):
                EPS.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[9]/table[2]/tbody/tr[1]/td[{i}]').text).replace(',',''))
                PER.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[9]/table[2]/tbody/tr[17]/td[{i}]').text).replace(',',''))
                EV_EBITDA.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[9]/table[2]/tbody/tr[29]/td[{i}]').text).replace(',',''))
                PSR.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[9]/table[2]/tbody/tr[26]/td[{i}]').text).replace(',',''))
                PBR.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[9]/table[2]/tbody/tr[17]/td[{i}]').text).replace(',',''))

            #수익성
            broswer.find_element_by_id("val_td1").click()
            time.sleep(0.5)
            #영업 이익률, 순이익률, ROE
            operating_profit_ratio = []
            net_profit_rate = []
            ROE = []
            for i in range(2,7):
                operating_profit_ratio.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[4]/td[{i}]').text).replace(',',''))
                net_profit_rate.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[7]/td[{i}]').text).replace(',',''))
                ROE.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[13]/td[{i}]').text).replace(',',''))

            #안전성
            #부채비율, 유동부채 비율, 순 부채 비율, 유동 비율, 당좌비율
            broswer.find_element_by_id("val_td3").click()
            time.sleep(0.5)
            debt_ratio = []
            current_liability_ratio = []
            net_debt_ratio = []
            current_rate = []
            current_account_ratio = []
            for i in range(2,7):
                debt_ratio.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[1]/td[{i}]').text).replace(',',''))
                current_liability_ratio.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[4]/td[{i}]').text).replace(',',''))
                net_debt_ratio.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[10]/td[{i}]').text).replace(',',''))
                current_rate.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[13]/td[{i}]').text).replace(',',''))
                current_account_ratio.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[16]/td[{i}]').text).replace(',',''))
            #성장성
            #매출액 증가율, 영업 이익 증가율, 순이익 증가율
            broswer.find_element_by_id("val_td2").click()
            time.sleep(0.5)
            sales_growth_rate = []
            operating_profit_growth_rate = []
            net_profit_growth_rate = []
            for i in range(2,7):
                sales_growth_rate.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[1]/td[{i}]').text).replace(',',''))
                operating_profit_growth_rate.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[4]/td[{i}]').text).replace(',',''))
                net_profit_growth_rate.append((broswer.find_element_by_xpath(f'/html/body/div/form/div[1]/div/div[2]/div[3]/div/div/div[5]/table[2]/tbody/tr[7]/td[{i}]').text).replace(',',''))
            

            #평가기준
            score = 0
            score_criteria2 = 0 # 몇년간 상승했는지 체크
            #순 부채 비율이 30%이하 인지
            try:
                for i in range(4):
                    if float(net_debt_ratio[i+1]) < 30:
                        score_criteria2 = score_criteria2 + 1
                    if score_criteria2 > 3:
                        score = score + 3
                        score_criteria2 = 0
            except:
                pass

            #투자활동 활동과 재무활동이 모두 + 인지
            try:
                for i in range(4):
                    if float(investment_activities[i+1]) > 0 and float(financial_activities[i+1]) > 0:
                        score_criteria2 = score_criteria2 + 1
                    if score_criteria2 >= 3:
                        score = score - 3
                        score_criteria2 = 0
            except:
                pass

            # EPS와 ROE가 꾸준히 오르는가
            try:
                for i in range(4):
                    if float(EPS[i]) < float(EPS[i+1]) or float(ROE[i]) < float(ROE[i+1]):
                        score_criteria2 = score_criteria2 + 1
                    if score_criteria2 > 3:
                        score = score + 3
                        score_criteria2 = 0
            except:
                pass

            # 유동비율 ,당좌비율이 100%가 넘고 부채비율이 200%인지
            try:
                for i in range(4):
                    if (float(current_liability_ratio[i+1]) > 100 and float(current_account_ratio[i+1]) > 100) and float(debt_ratio[i+1]) < 200:
                        score_criteria2 = score_criteria2 + 1
                    if score_criteria2 > 3:
                        score = score + 3
                        score_criteria2 = 0
            except:
                pass
            
            # 영업이익률이, 매출이익률,당기이익률 계속 오르는지
            try:
                for i in range(4):
                    if (float(sales_growth_rate[i+1]) > 0 and float(operating_profit_growth_rate[i+1]) > 0) or float(net_profit_growth_rate[i+1]) > 0:
                        score_criteria2 = score_criteria2 + 1
                    if score_criteria2 > 3:
                        score = score + 3
                        score_criteria2 = 0
            except:
                pass
            # 동종업계 평균보다 PER이 낮은가
            try:
                for i in range(4):
                    if float(PER[4]) < float(INDUSTRIES_AVG_PER):
                        score_criteria2 = score_criteria2 + 1
                    if score_criteria2 > 3:
                        score = score + 3
                        score_criteria2 = 0
            except:
                pass
            
            broswer.quit()
            print(stock_info[0],":",(score/15)*100)
        except:
            continue



